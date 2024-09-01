import  openpyxl
file_path='/Users/hwan/Library/Containers/com.tencent.xinWeChat/Data/Library/Application Support/com.tencent.xinWeChat/2.0b4.0.9/830da9def43de92ae65513342e7367b1/Message/MessageTemp/137f4c83ecf5057aec6d9038c0c6a9f9/File/产学研数据2.xlsx'
#导入结果
workbook = openpyxl.load_workbook(file_path)  # 加载已有文件
# 选择活动的工作表
sheet = workbook.active
# 获取最后一列的索引
last_col = sheet.max_column + 1
print(last_col)
sheet.cell(row=1,column=last_col,value='结果')
# 将列表中的元素写入最后一列
# for i, value in enumerate(results_list, start=1):
#     sheet.cell(row=i, column=last_col, value=value)
# 保存工作簿
workbook.save(file_path)
    # print(out)