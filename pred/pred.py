# -*- coding: GBK -*-
import json

import openpyxl
from keras_preprocessing.text import text_to_word_sequence

from data import build_corpus
from utils import load_model, extend_maps, prepocess_data_for_lstmcrf_test
import pandas as pd


# def pred():
train_word_lists=[]
train_tag_lists=[]
word2id={}
tag2id={}
bilstm_word2id={}
bilstm_tag2id={}
crf_word2id={}
crf_tag2id={}

# ����word2id, tag2id
train_word_lists, train_tag_lists, word2id, tag2id = \
    build_corpus("train")
bilstm_word2id, bilstm_tag2id = extend_maps(word2id, tag2id, for_crf=False)

crf_word2id, crf_tag2id = extend_maps(word2id, tag2id, for_crf=True)
bilstm_model = load_model('/Users/hwan/Downloads/named_entity_recognition-master/ckpts/bilstm_crf.pkl')
# bilstm_model = load_model('/Users/hwan/Downloads/named_entity_recognition-master/ckpts/bilstm.pkl')


# ��ȡExcel�ļ�
# file_path = '/Users/hwan/Library/Containers/com.tencent.xinWeChat/Data/Library/Application Support/com.tencent.xinWeChat/2.0b4.0.9/830da9def43de92ae65513342e7367b1/Message/MessageTemp/137f4c83ecf5057aec6d9038c0c6a9f9/File/��ѧ������1.xlsx'  # �滻Ϊ���ļ���·��
file_path='/Users/hwan/Library/Containers/com.tencent.xinWeChat/Data/Library/Application Support/com.tencent.xinWeChat/2.0b4.0.9/830da9def43de92ae65513342e7367b1/Message/MessageTemp/137f4c83ecf5057aec6d9038c0c6a9f9/File/��ѧ������2.xlsx'
df = pd.read_excel(file_path)

# ��ȡ�����е��ı��������б�
g_column_data = df['����'].tolist()
results_list=[]
# ������
# print(g_column_data)
for text in g_column_data:
    # ��txt����test_word_list
    # text=''
    test_word_lists=[]
    # text = u'5��24�����磬�廪��ѧ����Ժ���Ƚ�����Ժ��һ����Ժ������Ⱥ���ڵĴ���������������ѧ�ι۵��У����뱱����ѧ��ѧ�о������Ƚ������о�Ժ���Ƽ����������¹��ƽ���칫�ҵȲ��ſ�չѧϰ�᳹ϰ��ƽ��ʱ���й���ɫ�������˼�����������ѧ�������ι۵�������ѧ��ڲ�ƽ������У�����У��ɱ�����в�����л���������֡�������в��������浳֧������Ŵ�Ⱥ������������ǿ���Ƚ�Ժ��Ժ����Ĭ�����Ƽ�������������֣���࣬�¹��ư����μ���в�������Τ��廪��ѧ����Ժ��Ժ����ˮ�塢���������Ƚ�����Ժ��Ժ�����Ʒ��30������Աͬ־�μӻ����У���й�����Ա�ι��˱�����У��������У������ίԱ��칫����������ʦ�Ĵ����£������Ⱥ�ι�������¥��������ѧ¥��ѧ������5��¥����ʦ��Ԣ���������¥���ڲι۽������˽ⱱ����У���Ľṹ���ֺ͹��ܶ�λ����ҷ׷ױ�ʾ��������У���������ɡ��豸�Ƚ��������뱸�����ֳ�������ѧ��������չ�����Ǻ��¹��ƽ�������ϵ���ͷ���ι۱�����У���ι�֮����У���й�����Ա�ڼ����ѧԺ�������ԡ���չ����������֯���У������ˮƽ�Ƽ�������ǿ��Ϊ������н�����̸����̸���ֳ���̸���ϣ��Ŵ�Ⱥ���������Ա����ѧϰ��ϰ��ƽ����ǹ��ڸ�У�Ƽ����¹�������Ҫ����������ϰ��ƽ����ǿ��챱���廪ʱ����Ҫ��������ͬ־�Ǳ�ʾ��ϰ��ƽ����ǶԼ�ǿ��У�Ƽ����¹�����һϵ����Ҫ�������ݽ�겡����춨�򣬶Ը�ˮƽ�о��ʹ�ѧ��ȷ�µĶ�λ������µ�Ҫ�󡣴�ҽ��μ�������ڿ��챱����ѧ���廪��ѧʱ���������У���ְ�ѧ��ȷ���η��򣬽�������ʽ�ʦ���飬�γɸ�ˮƽ�˲�������ϵ������ԭʼ�����������ӿ��ƽ�����֯���У�����ǿ������ս�ԿƼ��������裬�Ը�ˮƽ�Ƽ����·����������չ��Ϊ�ӿ콨������Ƽ�ǿ����ʵ���й�ʽ�ִ��������µĹ��ס�Τ���ˮ�塢����ǿ�ֱ��������ؿ��й����������̸�ᷢ�ԣ������ҡ����ϵ�������Ϊ���Ŵ�Ⱥ��Τ���ˮ�塢����ǿ��Τ������˱����¹��Ʒ�չ���¿ռ佨�裬�Լ�����-��ƽ�����ںϴ������ĵĽ��������������У���Ǳ���չ�¹��Ƶ�����أ����¹��ƽ�ѧ����Ϊ�����ƽ�����ѧ���о��������¹����ԡ���������������ұ�����Թ��ٶࡱΪ�����ں������������б�����У�������ṩ���¿ռ�����Ļ���������ǿ�ص�ѧ�ƺ��ص㷽��������֣�ͬʱ����̽����Ч�Ĳ�ѧ��Э֧ͬ�Ż��ƣ���������ԭʼ���º����˼����Ĵ��´�ҵ��չ���ƶ�ѧ�������������Ͳ�ҵ���߶��ںϣ��ᶨ���Ƶ������ߺ��ۺ��Դ�ѧ��չ�¹��Ƶ���ɫ֮·����ˮ���ص�������廪��ѧ��ǿ����֯���С��������ƽ�2030�����ж��ƻ����йع������������ʾ���廪��ѧʼ�հѷ��������Ϊ���׷�󣬰ѹ��˹���ؽ�������ش�Ƽ�������Ϊ����ʹ��������������Ƽ�������ǿ��һ��ǿ�����⵼��ǣ��������֯���У�ͨ����չ2030�����ж��ƻ��ȣ��ڼ�ǿ�����о��͹ؼ����ļ������ء��ƽ���ѧ������ںϡ��ٽ��Ƽ����ź����ȷ��沼�֣�������������֯����ģʽ��������Դ���Ϻͼ������ƣ�ͳ����չУ������Դ�����ఴ�衢��׼֧�֣�Ϊ�ش������ṩ�������Ƽ����˲�Эͬ�Ļ������ϣ����Ǽ�ֿ��Ŵ��£���չ���ʿƼ�������ǣͷ�������ʴ��ѧ�ƻ������ˮƽ����ѧ������ƽ̨���������ʿƼ���������������ǿ�����˱�����ѧ�Ƽ����������������ѧ�᳹��ʵ���ĸ�����ս�Բ��𣬼�ָ�ˮƽ����̽�����ƶ�����������֯���У�����ǿ������ս�ԿƼ�����������ʵ�ָ�ˮƽ�Ƽ�������ǿ������ϸ�����˱�����ѧ���ƶ������о����е��ش��������ƿƼ�����ƽ̨��ϵ�������ش�ɹ�����������ʿƼ��������Ż����߱�����ϵ����Ĳ���ͳ�Ч��չʾ�˱�����ѧ�ڸ������ƶ��Ƽ����¹�����չ��ȫ��λ�������ս�ԿƼ���������ľٴ�ͳɼ�������Ⱥ��ʾ����ˮƽ�о��ʹ�ѧӦ���˲������������о���������������з��ӻ������á��廪��ѧ����Щ�����ǿ���֣�������չ�߲���˲������ƻ��������ж��ƻ���ȫ��ս�Լƻ����´�����ǿ�������о��Ϳ�չ����֯���С���������У���й������Ž���Эͬ���ƣ���ͬ�������ҿƼ�������ǿ��л��ָ����������ѧ��ָ�ˮƽ����̽���͸���������֯�������ϣ���Ժϵ��������ƶ�����̽����ͬʱ�Ի��ء�ƽ̨Ϊץ��ǿ������֯���С���������У���й������Ž����к���������̬����Ч��������ѧϰ�˴˵����澭�飬��ַ��Ӹ������ƺ���ɫ��Я�ִٽ���ˮƽ�Ƽ����·����������չ������Ⱥ��л�����ܽᷢ�Դ˴���ѧ�������������У���й������Ŷ������ǳ��������ѧϰ��ϰ��ƽ��ʱ���й���ɫ�������˼���ϰ��ƽ�������Ҫ�������񣬰��ո�������չ������ں����ֿ�չ�˿��й�������ҵ��̽�֣��ڼ�ǿ��ˮƽ�о��ʹ�ѧ��չ����֯���С���������ش�����ȷ��滥��ѧϰ������㾭���������˫����ʾ�����һ��Ҫ������ǿ�������������Я�ֹ�������������ǿ�����³Կࡢ���ڵ����İ��ﾫ��Ϊ�������ս�ԿƼ�����������Ƽ�ǿ���Ĺ�ͬĿ�겻�Ϸܶ������Ӱ'
    # text=u'2017��12��12�գ�Ϊ����ѧϰ����ʮ�Ŵ��񣬱�����ѧ�����뻷��ѧԺ2016��˶ʿ����֧�����廪��ѧ����ϵ������֧����չ������Ϊ������ǿ���ǿ ����������ˡ��������������ᵳԱ���ǰ�ڶ�ʮ�Ŵ󱨸��ѧϰ���ԡ�ר���������ʮ�Ŵ󱨸�֪ʶ��������ʫ�����С����������ۡ����������뵳�Ĵʡ��ȶ�����ʽ��һ����ǿ�˶�ʮ�Ŵ������ʱ�������˵�ʹ���������������ʶ�����Ӱ�ڡ�ר����������ڣ�����ǻ�ѧԺ����������������Ϊ������ѧϰʮ�Ŵ��� �����ֺ���ר�����ꡱ��������������������רҵ����ʮ�Ŵ���ָ���ġ�������̬�������л�����������չ��ǧ���ƣ����������ͼ�����ˮ��ɽ���ǽ�ɽ��ɽ��������������Լ����ĵ���ᡣ����ͬѧ�ǻ��㣺������飬������·������Ҫ�ߵ�������ء��ߵ������㡢�ߵ�����ͷ��ȥ�����й��Ĺ��顢��������飻�����Լ���ר��������רҵ֪ʶ���������⣬ͬʱ�����¡������е��������ѧУ�о���Ŭ������ѧ�����á���ѧ�н��ܽ�ϣ����õط��������ᷢչ���廪��ѧ����ϵ���������ˡ�ʢ���й� �ܶ��ഺ�����������������ϸ�������й��ڽ�������ȡ�õĻԻͳɾͣ����͡�һ��һ·����������������ܡ���ͬʱ����ͬѧ��־���Զ����̤ʵ�أ�Ҫ�������������й��ν��ܽ����������ʵ���б����ڷܺ�ѧ��̤ʵ�ϸɵ�̬�ȣ�һ��һ����ӡ��Ϊ�й�ʵ����������ִ���ǿ�������Լ���������ר������ڡ�ʮ�Ŵ󱨸�֪ʶ���������ڣ��ڳ��ĵ�Ա������ֳ�5�飬ÿ����б���ͬѧҲ���廪ͬѧ��֪ʶ������Ϊ����ͱش��������ڡ������������̼ȼ����ֲ���Ȥζ��ͬѧ�ǳ���Ӧ�ԣ�ʱ���������ء�������˼��ʱ��ϲЦ�տ����뵽�𰸡���������׷�����յ�һ��ͬѧ����ն������ùھ���ʮ�Ŵ󱨸�֪ʶ�����ڡ�ʫ�����С����ڣ�����ǻ�ѧԺ�����ú��廪����ϵ��־Զ���������ˡ����̵��������ǽ������������Ϊһ�壬��������߷���������˵�Ŷ������ֿ����������̵���������Ⱦ���ڳ���ÿһ���ˣ���ͬѧ�ǻع�������ķܶ�ʷ��Ҳ���������������δ����������ʱ���ͳ���ʱ���߿������У�ͬѧ�ǲ�ʱ���������ҵ�������ʫ�������ڡ��������ۡ����ڣ�����֧��ͬѧ�ԡ�����ʮ�Ŵ󱨸���ָ���ġ�����ǿ���ǿ ����������ˡ�̸̸����ΪӦ������������굳Ա�ļҹ��黳�����ε�������������ϼ����ߵ����ۣ�̸̸���۶�����Ӱ�족�����֧�������������飬̸̸��θ��õ��ƽ���У���㵳��֯���裬�Ӷ�������֧��������ȷ��ѧ����֧����������ѧ����ɫ�����ܷ��ӵ��Ļ��㱤�����á��͡�̸̸���ʮ�Ŵ󱨸��й�����̬������������⣬�Լ��о�����Ա��ƽʱ�Ŀ���ѧϰ��Ӧ����ƶ���̬�������衱Ϊ���������������ۡ��е�Ա̸�������굳ԱӦ������л��Ļ�����������ͬ���ڴ˻����ϣ����ҹ��黳�����ε������뵽����ѧϰ���е���ȥ�����Լ��ķ�չ���������ķ�չ���ܽ����һ���е�Ա��ʾ�ں�����Ϣ����ʱ�������������������ɡ��������ʵ�������ͬʱ������ѧ����ԱҲҪ�б��������۵��������е�Ա�Ա�����У��֧����չ���Ĳ�ͬ����ܽᾭ�飬̽���Ե����������ѧ��ʵ�ʵĵ�����ʽ���е�Ա���ʮ�Ŵ󱨸���������о�������ʾ�о�����ԱҪ������ѧϰ�Ϳ���ʵ��ͳ������������Ҫ�����������߾��⣬���ܽ����гɹ�ת����Ϊ�й�����̬�����������ɫ��չ֮·�ṩ�Ƽ�֧�š����������ڡ������뵳�Ĵʡ����ڣ���Ա��ȫ��������������ȭ��������ׯ�����ġ���Ա��һ��һ�䣬����������ᶨ���ơ��ڵ���ǰ��������һ�α���������Ա��Զ�������ģ�ʼ���μ�ʹ���������뵳�Ĵʱ���֧������������һ���ḻ�˵��ջ����ʽ�������������������֧����Ա��ʮ�Ŵ�����й��ɾ͵��˽⣻��ʮ�Ŵ�֪ʶ��������֧����Ա�Ƕ�ʮ�Ŵ󱨸���Ȼ���ģ���ʫ�����С���֧����Ա�ǻع������ʷ��չ������δ�������������ۡ�������֧����Ա�ǵĻ����Ժʹ����ԣ��������뵳�Ĵʡ���ͬѧ���μǳ��ģ�����ʹ����ͬʱ�����λҲ��ͬѧ���������ʶ����Ϊ���굳Ա�������뵣�����׷ױ�ʾҪ������֪��ѧ�����õػ�����ᣬΪ�й�ʵ����������ִ���ǿ������������ר�����ӣ�ѧϰ�᳹ʮ�Ŵ���༭��ɽʯ'

    # print(text)
    text = ' '.join(text)
    token_h = text_to_word_sequence(text)
    test_word_lists.append(token_h)
    # print(test_word_lists)
    # info='Ԥ���㷨Ϊ��'
    out=''



    bilstm_model.model.bilstm.bilstm.flatten_parameters()  # remove warning
    test_word_lists= prepocess_data_for_lstmcrf_test(
        test_word_lists, test=True
    )
    pred_list = bilstm_model.test2(test_word_lists,crf_word2id, crf_tag2id)

    #

    #
    # bilstm_model.model.bilstm.flatten_parameters()  # remove warning
    # # test2ֱ��Ԥ��
    # pred_list= bilstm_model.test2(test_word_lists,bilstm_word2id, bilstm_tag2id)

    # print(pred_list)
    for i in range(len(pred_list[0])):
        out_i=test_word_lists[0][i]+" "+pred_list[0][i]+'\n'
        out+=out_i
    # self.out_text.set(out)

    result={
        'DATE':[],
        'SCH': [],
        'COOP': [],
        'CXY': [],
        'JP': []
    }
    i = 0
    while i<len(pred_list[0]):
        #��������һ��'-'ʱ������'-'ʱ����
        while len(pred_list[0][i])>1 :
            added=''
            #���һ��
            # if i==len(pred_list[0]):
            #     if len(pred_list[0][i])>1:
            #         added+=test_word_lists[0][i]
            # else:
            try:
                while len(pred_list[0][i+1])>1:
                    added+=test_word_lists[0][i]
                    i+=1
                    if i == len(pred_list[0])-1:
                        break

                added+=test_word_lists[0][i]
            except IndexError as e:
                print(i)
                print(len(pred_list[0]))
            #��һ��Ϊ0�����һ��
            if len(pred_list[0][i])>1 and len(added)>1:
                if pred_list[0][i][2:] == 'DATE':
                    result['DATE'].append(added)
                if pred_list[0][i][2:] == 'SCH':
                    result['SCH'].append(added)
                if pred_list[0][i][2:] == 'COOP':
                    result['COOP'].append(added)
                if pred_list[0][i][2:] == 'CXY':
                    result['CXY'].append(added)
                if pred_list[0][i][2:] == 'JP':
                    result['JP'].append(added)
            i+=1
            if i == len(pred_list[0]):
                break
        i+=1

    # print(result)
    results_list.append(result)
    # print(results_list)
print(results_list)
print(len(results_list))

# # ����Ϊ JSON �ļ�
# with open('result2.json', 'w') as f:
#     json.dump(results_list, f, indent=4)

#������
workbook = openpyxl.load_workbook(file_path)  # ���������ļ�
# ѡ���Ĺ�����
sheet = workbook.active
# ��ȡ���һ�е�����
last_col = sheet.max_column + 1
sheet.cell(row=1,column=last_col,value='ʱ��')
sheet.cell(row=1,column=last_col+1,value='ѧУ')
sheet.cell(row=1,column=last_col+2,value='������')
sheet.cell(row=1,column=last_col+3,value='��ѧ�л���')
sheet.cell(row=1,column=last_col+4,value='������')
# ���б��е�Ԫ��д�����һ��
for i, value in enumerate(results_list, start=2):
    sheet.cell(row=i, column=last_col, value=str(value['DATE']))
    sheet.cell(row=i, column=last_col+1, value=str(value['SCH']))
    sheet.cell(row=i, column=last_col+2, value=str(value['COOP']))
    sheet.cell(row=i, column=last_col+3, value=str(value['CXY']))
    sheet.cell(row=i, column=last_col+4, value=str(value['JP']))


# ���湤����
workbook.save(file_path)
    # print(out)